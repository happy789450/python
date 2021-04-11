import datetime

from openpyxl import Workbook

wb = Workbook()
ws = wb.active


ws['A2'] = 1
# ws.append([1, 2, 3])
# ws.append([2, 2, 3])
# ws['A2'] = datetime.datetime.now()
# ws.append([datetime.datetime.now()])
# ws['A3'] = 666
# ws['A5'] = 'aaa'
wb.save("/home/rice/桌面/test/sample.xlsx")